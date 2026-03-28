"""
main.py — Run the Vet School MILP Timetabler and save results
=============================================================
Usage:
    python main.py                  # weeks 9-10 (default)
    python main.py --start-week 15  # weeks 15-16

Outputs to proposedTimetable/:
    solution_weeks_<s>_<e>.xlsx     — flat event list (all sources, enriched metadata)
    timetable_grid_weeks_<s>_<e>.xlsx — pivot grid: one sheet per day, rows=rooms, cols=hours
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from timetabler import Timetabler
from utils import DAY_ORDER, HOUR_ORDER, parse_timeslot

OUT_DIR = Path("proposedTimetable")

# Colour palette for Source column in the grid
_SOURCE_FILL = {
    "milp":        PatternFill("solid", fgColor="C6EFCE"),   # green
    "fixed_vet":   PatternFill("solid", fgColor="DDEBF7"),   # blue
    "fixed_non_vet": PatternFill("solid", fgColor="FCE4D6"), # orange
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _enrich(sol: pd.DataFrame, events_df: pd.DataFrame) -> pd.DataFrame:
    """Left-join event metadata (name, module, type) onto the solution rows."""
    want = ["Event ID", "Event Name", "Event Type", "Module Code", "Module Name"]
    available = [c for c in want if c in events_df.columns]
    meta = events_df[available].drop_duplicates(subset="Event ID")
    return sol.merge(meta, on="Event ID", how="left")


def _cell_label(row: pd.Series) -> str:
    """Short display string for a timetable grid cell."""
    parts = [str(row["Event ID"])]
    if pd.notna(row.get("Event Name")):
        parts.append(str(row["Event Name"]))
    if pd.notna(row.get("Module Code")):
        parts.append(f"[{row['Module Code']}]")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_flat_solution(sol: pd.DataFrame, path: Path) -> None:
    """Write the enriched flat solution DataFrame to Excel."""
    col_order = [
        "Event ID", "Event Name", "Event Type", "Module Code", "Module Name",
        "Room", "Timeslot", "Source", "Event Size", "Room Capacity",
        "Duration (minutes)", "Weeks", "Semester",
    ]
    cols = [c for c in col_order if c in sol.columns]
    sol[cols].to_excel(path, index=False)
    print(f"  Flat solution  → {path}  ({len(sol):,} rows)")


def write_timetable_grid(sol: pd.DataFrame, path: Path) -> None:
    """
    Write a pivot timetable to Excel: one sheet per day.
    Rows = rooms, columns = hour slots, cell = Event ID + name.
    Only milp and fixed_vet rows are shown (vet-room events only).
    """
    grid = sol[sol["Source"].isin(["milp", "fixed_vet"])].copy()
    grid["Day"] = grid["Timeslot"].apply(lambda ts: parse_timeslot(ts)[0])
    grid["Hour"] = grid["Timeslot"].apply(lambda ts: parse_timeslot(ts)[1])
    grid["Label"] = grid.apply(_cell_label, axis=1)

    active_days = [d for d in DAY_ORDER if d in grid["Day"].values]
    active_hours = [h for h in HOUR_ORDER if h in grid["Hour"].values]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for day in active_days:
            day_df = grid[grid["Day"] == day]
            if day_df.empty:
                continue

            pivot = day_df.pivot_table(
                index="Room",
                columns="Hour",
                values="Label",
                aggfunc=lambda xs: " | ".join(str(v) for v in xs),
            )
            pivot = pivot.reindex(columns=[h for h in active_hours if h in pivot.columns])
            pivot.to_excel(writer, sheet_name=day)

            # Light formatting
            ws = writer.sheets[day]
            for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row):
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = 28
            for row_cells in ws.iter_rows(min_row=2):
                for cell in row_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[1].height = 18
            for cell in ws[1]:
                cell.font = Font(bold=True)

    print(f"  Timetable grid → {path}  ({len(active_days)} day sheet(s))")


def write_summary_sheet(sol: pd.DataFrame, t: Timetabler, path: Path) -> None:
    """Write a one-page summary with key statistics."""
    milp_rows = sol[sol["Source"] == "milp"]
    room_conflicts = milp_rows[["Room", "Timeslot"]].duplicated().sum()

    rows = [
        ("Weeks covered",        sorted(t.target_weeks)),
        ("Free events (MILP)",   len(t.E)),
        ("MILP-assigned",        int((sol["Source"] == "milp").sum())),
        ("Fixed-vet",            int((sol["Source"] == "fixed_vet").sum())),
        ("Fixed-non-vet",        int((sol["Source"] == "fixed_non_vet").sum())),
        ("Vet rooms available",  len(t.R)),
        ("Timeslots available",  len(t.T)),
        ("Unique rooms used",    int(sol["Room"].nunique())),
        ("Unique timeslots used",int(sol["Timeslot"].nunique())),
        ("Room conflicts (MILP)",int(room_conflicts)),
        ("Solve status",         t._solve_status or "unknown"),
        ("Warnings",             len(t.warnings)),
    ]
    df = pd.DataFrame(rows, columns=["Metric", "Value"])
    df.to_excel(path, index=False)
    print(f"  Summary        → {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Run the Vet School MILP timetabler and save results."
    )
    parser.add_argument(
        "--start-week", type=int, default=9,
        help="First week of the 2-week window (default: 9)",
    )
    args = parser.parse_args()

    start = args.start_week
    end   = start + 1          # n_weeks = 2

    # --- Solve ---
    t = Timetabler(start_week=start, n_weeks=2)
    status = t.run()
    t.summary()

    # --- Prepare output directory ---
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    tag = f"weeks_{start}_{end}"

    if status != "optimal":
        print(f"\nSolver returned '{status}' — no solution to write.")
        unassigned = t.get_unassigned_events()
        if unassigned:
            print(f"Unassigned events ({len(unassigned)}): {unassigned[:20]}")
        sys.exit(1)

    # --- Build and enrich solution ---
    sol = t.get_solution()
    if t.events is not None:
        sol = _enrich(sol, t.events)

    # --- Sanity check ---
    milp = sol[sol["Source"] == "milp"]
    conflicts = milp[["Room", "Timeslot"]].duplicated().sum()
    if conflicts:
        print(f"\nWARNING: {conflicts} (Room, Timeslot) duplicates in MILP rows!")
    else:
        print("\nNo room conflicts in MILP solution ✓")

    # --- Write outputs ---
    print(f"\nWriting outputs to {OUT_DIR}/")
    write_flat_solution(sol,  OUT_DIR / f"solution_{tag}.xlsx")
    write_timetable_grid(sol, OUT_DIR / f"timetable_grid_{tag}.xlsx")
    write_summary_sheet(sol, t, OUT_DIR / f"summary_{tag}.xlsx")

    print(f"\nDone. {len(sol):,} events written ({len(milp):,} MILP-assigned).")


if __name__ == "__main__":
    main()
